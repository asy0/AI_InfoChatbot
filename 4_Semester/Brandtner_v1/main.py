import os
import pickle
import sys
import csv
from sentence_transformers import SentenceTransformer, util
import torch
from pdf_reader import read_pdf
from nlp_processing import extract_sentences, embed_sentences, group_sentences, generate_paraphrases, filter_best_paraphrases, find_best_match
from cross_encoder_reranker import rerank_with_cross_encoder
from nlp_processing import expand_query
import re
import numpy as np
from itertools import chain
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl import load_workbook
import gc


def analyze_pdfs(folder_path):
    pdf_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith('.pdf')]
    if not pdf_files:
        print("Keine PDFs im Ordner gefunden.")
        return

    cache_path = "pickle_cache/sentence_cache.pkl"
    all_sentences = load_embeddings_from_cache(cache_path)

    if not all_sentences:
        print("\n Keine gecachten Embeddings gefunden – verarbeite PDFs...")
        all_sentences = []

        for pdf_path in pdf_files:
            print(f"\nAnalysiere: {os.path.basename(pdf_path)}")

            try:
                raw_text = read_pdf(pdf_path)
                print("\nVollständige Sätze/Paragraphen:")

                sentences = extract_sentences(raw_text)
                grouped_chunks = group_sentences(sentences, group_size=1)
                all_sentences.extend(embed_sentences(grouped_chunks, os.path.basename(pdf_path)))

            except Exception as e:
                print(f"Fehler bei der Verarbeitung von {os.path.basename(pdf_path)}: {e}")
        
        # Cache embeddings
        save_embeddings_to_cache(all_sentences, cache_path)
    return all_sentences

def evaluate_mode(test_csv_path, all_sentences):
    output_path = "./answers/evaluated_answers.csv"
    model = SentenceTransformer("intfloat/multilingual-e5-large", device="cuda")

    questions = []
    references = []

    # Read CSV
    with open(test_csv_path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            questions.append(row['frage'])
            references.append(row['referenztext'])

    results = []
    for idx, (query, reference) in enumerate(zip(questions, references)):
        print(f"Frage {idx+1}: {query}")

        # Use chatbot to generate answer
        raw_paraphrases = generate_paraphrases(query, num_return_sequences=5)
        paraphrases = [query] + filter_best_paraphrases(query, raw_paraphrases, top_n=8)
        all_matches = []
        all_matches = []
        for variant in paraphrases:
            expanded = expand_query(variant)
            bi_matches = find_best_match(expanded, all_sentences)
            bi_matches_raw = [m[0] for m in bi_matches]  # Extract (sentence, embedding, source)
            cross_matches = rerank_with_cross_encoder(expanded, bi_matches_raw, top_k=10)
            all_matches.extend(cross_matches)

        matches = sorted(all_matches, key=lambda x: x[1], reverse=True)
        seen = set()
        unique_matches = []
        for match in matches:
            sentence = match[0][0]
            if sentence not in seen:
                seen.add(sentence)
                unique_matches.append(match)
        matches = [m for m in unique_matches if m[1] > 0.1]
        not_matches = [m for m in unique_matches if m[1] <= 0.1]

        if matches:
            bot_answer = matches[0][0][0]
        else:
            bot_answer = not_matches[0][0][0]
            print("Keine passenden Antworten gefunden, nehme Fallback.")

        embeddings = model.encode(
            [f"query: {bot_answer}", f"passage: {reference}"],
            convert_to_tensor=True
        )

        # Cosine similarity (as before)
        similarity_score = float(util.cos_sim(embeddings[0], embeddings[1])[0][0])

        # Optional: L2 distance (lower is better)
        l2_distance = float(np.linalg.norm(embeddings[0].cpu().numpy() - embeddings[1].cpu().numpy()))

        results.append({
            "index": idx + 1,
            "question": query,
            "reference": reference,
            "myanswer": bot_answer,
            "score": round(similarity_score, 4),
            "l2_distance": round(l2_distance, 4),
        })
        torch.cuda.empty_cache()
        gc.collect()

    # Save as Excel
    df = pd.DataFrame(results)
    output_path = "./answers/evaluated_answers.xlsx"
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    score_col = "E"  # Assuming 'score' is column E
    start_row = 2
    end_row = len(df) + 1
    cell_range = f"{score_col}{start_row}:{score_col}{end_row}"

    # RED fill < 0.7
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='lessThan', formula=['0.7'], fill=red_fill))

    # YELLOW fill 0.7–0.85
    yellow_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='between', formula=['0.7', '0.8499'], fill=yellow_fill))

    # GREEN fill ≥ 0.85
    green_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator='greaterThanOrEqual', formula=['0.85'], fill=green_fill))
    # Save workbook
    wb.save(output_path)
    print(f"Scores gespeichert in {output_path}")

def save_embeddings_to_cache(data, cache_path="pickle_cache/sentence_cache.pkl"):
    os.makedirs(os.path.dirname(cache_path), exist_ok=True)
    with open(cache_path, "wb") as f:
        pickle.dump(data, f)
    print(f" Embeddings gespeichert in {cache_path}")


def load_embeddings_from_cache(cache_path="pickle_cache/sentence_cache.pkl"):
    try:
        with open(cache_path, "rb") as f:
            data = pickle.load(f)
        print(f" Geladene gecachte Embeddings aus {cache_path}")
        return data
    except FileNotFoundError:
        return None

def split_bullets(sentences):
    return list(chain.from_iterable(re.split(r"[•\-]", s) if any(b in s for b in "•-") else [s] for s in sentences))


if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    folder_path = os.path.join(script_dir, "data")
    all_sentences = analyze_pdfs(folder_path)

    if not all_sentences:
        print("Keine Embeddings verfügbar.")
        sys.exit(1)

    args = sys.argv[1:]
    if args:
        arg_path = args[0]
        evaluate_mode(arg_path, all_sentences)
    else:
        # Loop for user queries
        while True:
            query = input("\nFrage an den Bot (oder 'exit'): ")
            if query.strip().lower() in ["exit", "quit"]:
                break

            raw_paraphrases = generate_paraphrases(query, num_return_sequences=5)
            paraphrases = [query] + filter_best_paraphrases(query, raw_paraphrases, top_n=8)

            all_matches = []
            for variant in paraphrases:
                expanded = expand_query(variant)
                variant_matches = find_best_match(expanded, all_sentences)
                all_matches.extend(variant_matches)

            matches = sorted(all_matches, key=lambda x: x[1], reverse=True)
            seen = set()
            unique_matches = []
            for match in matches:
                sentence = match[0][0]
                if sentence not in seen:
                    seen.add(sentence)
                    unique_matches.append(match)
            matches = unique_matches
            matches = [m for m in matches if m[1] > 0.1]

            if matches:
                best_match = matches[0]
                print("\n💡 Beste Antwort:")
                print(f"{best_match[0][0]}  [📄 {best_match[0][2]} | 🔍 Score: {best_match[1]:.2f}]")

                print("\n\n Weitere relevante Antworten:\n")
                for (sentence, _, source), score in matches[1:]:
                    print(f"- {sentence}  [📄 {source} | 🔍 Score: {score:.2f}]")
            else:
                print("Keine passenden Antworten gefunden.\n")
                print("Möglicherweise relevante Antwort (niedriger Score):")

                from sentence_transformers import util
                from sentence_transformers import SentenceTransformer
                sentence_model = SentenceTransformer("intfloat/multilingual-e5-large", device="cuda")
                query_embedding = sentence_model.encode([f"query: {expanded}"], convert_to_tensor=True, device="cuda")[0]

                scored = [
                    ((sentence, emb, source), float(util.cos_sim(query_embedding, emb)[0][0]))
                    for (sentence, emb, source) in all_sentences
                ]
                scored = sorted(scored, key=lambda x: x[1], reverse=True)

                if scored:
                    fallback = scored[0]
                    print(f"{fallback[0][0]}  [📄 {fallback[0][2]} | 🔍 Score: {fallback[1]:.2f}]")
                else:
                    print("Keine Antworten verfügbar.")